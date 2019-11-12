# -*- coding: utf-8 -*-
# @Time    : 19-1-11 上午9:33
# @Author  : Felix Wang

import xlrd
from openpyxl import load_workbook
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font


class OperationExcel(object):
    def __init__(self, filename):
        self.filename = filename
        try:
            self.__workbook = xlrd.open_workbook(self.filename)
        except Exception as e:
            self.creatwb(self.filename)
            self.__workbook = xlrd.open_workbook(self.filename)
        self.__wb = load_workbook(self.filename)
        self.ws = self.__wb.active

    def __enter__(self):
        """进入with语句的时候被调用"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """离开with的时候被with调用"""
        self.__wb.save(self.filename)
        self.__wb.close()

    def get_sheet_names(self):
        return self.__wb.sheetnames

    def get_data(self, row=1, col=1, sheet_name=None, flag='0'):
        """
        :param sheet_name: 表名
        :param row: 行位置
        :param col: 列位置
        :param flag:标志，0：表示获取某行某列，1：获取某行，2：获取某列
        :return:返回
        """
        if sheet_name is None:
            sheet_name = self.get_sheet_names()[0]
        sheet = self.__workbook.sheet_by_name(sheet_name)
        # c = sheet.ncols  # 最大列数
        # w = sheet.nrows  # 最大行数
        # print(c, w)
        try:
            if str(flag) == '0':
                return sheet.row_values(row - 1)[col - 1]
            if str(flag) == '1':
                return sheet.row_values(row - 1)
            if str(flag) == '2':
                return sheet.col_values(col - 1)
        except Exception as e:
            print('该位置没有数据', e)

    def get_all_data(self, sheet_name=None):
        if sheet_name is None:
            sheet_name = self.get_sheet_names()[0]
        sheet = self.__workbook.sheet_by_name(sheet_name)
        data = []
        for i in range(sheet.nrows):
            data.append(sheet.row_values(i))
        return data

    def get_max_row_and_col(self):
        return {'max_rows': self.ws.max_row, 'max_cols': self.ws.max_column}

    # 新建excel
    def creatwb(self, filename):
        wb = openpyxl.Workbook()
        wb.save(filename=filename)
        print("新建Excel：" + filename + "成功")

    # 写入数据
    def write_data(self, col, row, data):
        self.ws.cell(row, col, data)

    def write_all_data(self, data_list, left_offset=0, top_offset=0):
        """
        :param data_list: 列表或者元组[[],[],[]...],里面的每一个表示一行
        :param left_offset: 左边偏移
        :param top_offset: 上边偏移
        :return:
        """
        left_offset = left_offset if str(left_offset).isdigit() and left_offset > 0 else 0
        top_offset = top_offset if str(top_offset).isdigit() and top_offset > 0 else 0
        lens = len(data_list)
        for index, data in enumerate(data_list):
            for i, dd in enumerate(data):
                self.ws.cell(index + top_offset + 1, i + left_offset + 1, dd)
            print(index + 1, lens, data)

    def set_color(self, col, row, color):
        """
        设置单元格颜色
        :param col:
        :param row:
        :param color: 颜色：如 FFFFFF
        :return:
        """
        # 单元格填充颜色
        self.ws.cell(row, col).fill = PatternFill(fill_type='solid', fgColor=color)

    def set_hyperlink(self, col, row, hyperlink):
        """
        设置超链接
        :param col:
        :param row:
        :param hyperlink:超链接地址
        :return:
        """
        self.ws.cell(row, col).hyperlink = hyperlink

    def set_font(self, col, row, font):
        self.ws.cell(row, col).font = font

    # 往Excel中插入图片
    def insert_img(self, img_name, place=None):
        img = Image(img_name)
        if place is None:
            self.ws.add_image(img)
        else:
            self.ws.add_image(img, place)  # 例如'A1'


if __name__ == '__main__':
    with OperationExcel('a.xlsx') as op:
        d = op.get_sheet_names()
        print(d)
        # print(op.get_data('Sheet2', 1, 2, 0))
        # print(op.get_all_data('Sheet1'))
        op.write_data(1, 3, 'test')
        op.write_data(2, 3, 't2est')

        # op.insert_img(r'1234.jpg')
        # op.insert_img(r'1234.jpg', 'D3')
        op.write_all_data([[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12, 15]], top_offset=-10, left_offset=15)
        op.set_color(3, 5, 'ff0000')
        op.write_data(3, 5, 'hello')

        font = Font(
            name="宋体",
            size=15,
        )
        op.set_font(3, 5, font)

'''

　　####循环插入字典可以使用这种方法
row_and_col = op.get_max_row_and_col()
max_rows = row_and_col['max_rows']
max_cols = row_and_col['max_cols']
if max_rows == 1 and max_cols == 1:
    print('开始插入Excel标题')
    i = 1
    for key in data.keys():
        op.write_data( i, 1, key)
        op.write_data( i, 2, data[key])
        i += 1
    print('插入Excel标题成功')
if max_rows > 1 or max_cols > 1:
    print('开始插入Excel数据')
    j = 1
    for key in data.keys():
        op.write_data( j, max_rows + 1, data[key])
        j += 1
    print('插入Excel数据成功')
# 循环插入字典时
'''
